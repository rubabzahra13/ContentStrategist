#!/usr/bin/env python3
"""
Tests to confirm Excel files are actually written and not empty
"""

import unittest
import sys
import os
import tempfile
import shutil
from datetime import datetime

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from core.excel_exporter import export_to_excel
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Excel testing unavailable: {e}")
    EXCEL_AVAILABLE = False

class TestExcelGeneration(unittest.TestCase):
    """Tests for Excel file generation and content validation"""

    def setUp(self):
        """Set up test environment"""
        self.test_dir = tempfile.mkdtemp()
        self.sample_calendar_text = """
Day | Reel Title | Hook Script (0-2s) | Body Breakdown (3-20s) | Close/CTA (20-30s) | Format Style | Audio Style | Hashtag Strategy | Production Notes | Optimization Tips
Day 1 | "3 Tools Secretly Running Successful Businesses" | "Behind the scenes of elite entrepreneurs..." | "Tool 1: Advanced planning systems - outlines in 60s. Tool 2: Visual automation - automated visuals. Tool 3: Workflow systems - processes while you sleep" | "Save this before it's taken down" | Screen demo + talking head | Suspenseful tech beat | #businesstools #entrepreneur | Dramatic office clips, zoom effects | Bold text, capitalize SECRETLY
Day 2 | "This Single Prompt Transformed My Content Strategy" | "Watch this transformation happen live..." | "I'll show you the exact prompt, the input I gave it, and the mind-blowing output it produced" | "DM 'PROMPT' for the exact framework" | Voiceover + screen recording | Trending upbeat track | #contentcreation #businesshacks #entrepreneur | Screen recordings of actual process | Use progress bars, before/after
Day 3 | "Don't Start a Business in 2025 Without These Systems" | "If you're building anything this year..." | "Customer Research: Advanced research tools. Naming: Strategic branding frameworks. Funnels: Automated marketing systems" | "Which one blew your mind? 👀" | Slide carousel + graphics | Tech/futuristic beat | #business2025 #startup | Clean graphics, smooth transitions | Bullet points, checkmarks, progress
"""

    def tearDown(self):
        """Clean up test environment"""
        shutil.rmtree(self.test_dir, ignore_errors=True)

    @unittest.skipUnless(EXCEL_AVAILABLE, "Excel dependencies not available")
    def test_excel_file_creation(self):
        """Test that Excel file is actually created"""
        output_path = os.path.join(self.test_dir, "test_calendar.xlsx")
        
        # Export to Excel
        export_to_excel(self.sample_calendar_text, output_path)
        
        # Verify file exists
        self.assertTrue(os.path.exists(output_path), "Excel file was not created")
        
        # Verify file is not empty
        file_size = os.path.getsize(output_path)
        self.assertGreater(file_size, 1000, f"Excel file too small ({file_size} bytes)")

    @unittest.skipUnless(EXCEL_AVAILABLE, "Excel dependencies not available")
    def test_excel_content_validation(self):
        """Test that Excel file contains expected content"""
        output_path = os.path.join(self.test_dir, "test_calendar.xlsx")
        
        # Export to Excel
        export_to_excel(self.sample_calendar_text, output_path)
        
        # Load and validate content
        workbook = load_workbook(output_path)
        worksheet = workbook.active
        
        # Check header row
        expected_headers = [
            "Day", "Reel Title", "Hook Script (0-2s)", "Body Breakdown (3-20s)",
            "Close/CTA (20-30s)", "Format Style", "Audio Style", "Hashtag Strategy",
            "Production Notes", "Optimization Tips"
        ]
        
        header_row = [cell.value for cell in worksheet[1]]
        for i, expected_header in enumerate(expected_headers):
            self.assertEqual(header_row[i], expected_header, f"Header mismatch at column {i+1}")

        # Check data rows
        self.assertGreaterEqual(worksheet.max_row, 4, "Should have at least header + 3 data rows")
        
        # Verify first data row content
        first_data_row = [cell.value for cell in worksheet[2]]
        self.assertEqual(first_data_row[0], "Day 1")
        self.assertIn("3 AI Tools", first_data_row[1])
        self.assertIn("Behind the scenes", first_data_row[2])

    @unittest.skipUnless(EXCEL_AVAILABLE, "Excel dependencies not available")
    def test_excel_formatting(self):
        """Test that Excel file has proper formatting"""
        output_path = os.path.join(self.test_dir, "test_calendar.xlsx")
        
        # Export to Excel
        export_to_excel(self.sample_calendar_text, output_path)
        
        # Load and check formatting
        workbook = load_workbook(output_path)
        worksheet = workbook.active
        
        # Check header formatting
        header_cell = worksheet['A1']
        self.assertIsNotNone(header_cell.fill.start_color, "Header should have background color")
        self.assertTrue(header_cell.font.bold, "Header should be bold")
        
        # Check column widths are set
        self.assertGreater(worksheet.column_dimensions['B'].width, 20, "Title column should be wide")
        self.assertGreater(worksheet.column_dimensions['C'].width, 30, "Hook column should be wide")

    @unittest.skipUnless(EXCEL_AVAILABLE, "Excel dependencies not available")
    def test_malformed_input_handling(self):
        """Test Excel generation with malformed input"""
        malformed_inputs = [
            "Invalid data without pipes",
            "Day 1 | Only | Two | Columns",
            "Day 1|Too|Many|Columns|Here|More|Than|Expected|Extra|Data|Overflow|Problem",
            "",
        ]
        
        for i, malformed_input in enumerate(malformed_inputs):
            with self.subTest(input=malformed_input):
                output_path = os.path.join(self.test_dir, f"malformed_{i}.xlsx")
                
                try:
                    export_to_excel(malformed_input, output_path)
                    
                    if os.path.exists(output_path):
                        # File created - check it's valid
                        workbook = load_workbook(output_path)
                        worksheet = workbook.active
                        self.assertGreaterEqual(worksheet.max_row, 1, "Should have at least header row")
                    
                except Exception as e:
                    # Should handle gracefully, not crash
                    self.assertIsInstance(e, (ValueError, FileNotFoundError), 
                                        f"Should handle malformed input gracefully: {e}")

    @unittest.skipUnless(EXCEL_AVAILABLE, "Excel dependencies not available")
    def test_empty_calendar_handling(self):
        """Test handling of empty calendar content"""
        empty_inputs = [
            "",
            "   ",
            "Header only | with | pipes",
        ]
        
        for empty_input in empty_inputs:
            with self.subTest(input=empty_input):
                output_path = os.path.join(self.test_dir, "empty_test.xlsx")
                
                # Should either create valid file or raise appropriate error
                try:
                    export_to_excel(empty_input, output_path)
                    if os.path.exists(output_path):
                        # If file created, should have minimum structure
                        workbook = load_workbook(output_path)
                        worksheet = workbook.active
                        self.assertGreaterEqual(worksheet.max_row, 1)
                except ValueError:
                    # Acceptable to reject empty input with ValueError
                    pass

if __name__ == '__main__':
    print("📊 RUNNING EXCEL GENERATION TESTS")
    print("=" * 50)
    
    if EXCEL_AVAILABLE:
        unittest.main(verbosity=2, exit=False)
        print("\n✅ EXCEL TESTS COMPLETED")
    else:
        print("❌ Excel dependencies not available - install openpyxl to run tests")
        print("Run: pip install openpyxl")