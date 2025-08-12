# === FILE: core/excel_exporter.py ===
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def export_to_excel(table_text, filename, include_transcripts=False):
    """Export calendar data to Excel with proper formatting and error handling"""

    try:
        # Split text into lines and filter for data rows
        lines = table_text.strip().split("\n")
        data_lines = [line.strip() for line in lines if "|" in line]

        print(f"📝 Processing {len(lines)} total lines, {len(data_lines)} contain '|'")

        # Detect if transcripts are included by checking for "Transcript" in any line
        has_transcripts = any("Transcript" in line for line in data_lines)
        if has_transcripts:
            include_transcripts = True
            print("🎬 Detected transcript column in data")

        # Find header row and data rows
        header_row = None
        data_rows = []

        # Determine expected column count
        expected_columns = 11 if include_transcripts else 10

        for i, line in enumerate(data_lines):
            # Look for header row - more flexible detection
            if i == 0 or any(keyword in line.lower() for keyword in ["day", "date", "title", "hook", "content", "body", "cta"]):
                if not header_row and "|" in line:
                    header_row = [cell.strip() for cell in line.split("|") if cell.strip()]
                    expected_columns = len(header_row)
                    print(f"📋 Detected header with {expected_columns} columns: {header_row}")
                    continue

            # Process data rows - more flexible detection
            if "|" in line and line.strip():
                # Split and clean each row
                row = [cell.strip() for cell in line.split("|")]
                # Remove empty cells at the beginning and end
                while row and not row[0]:
                    row.pop(0)
                while row and not row[-1]:
                    row.pop()
                
                # Skip if it's clearly a header row we missed (but we already found the header)
                if header_row and row and any(keyword in row[0].lower() for keyword in ["day", "date", "title", "hook", "content", "body", "cta"]):
                    # Check if this is actually a header or data
                    if not (row[0].lower().startswith('day ') or row[0].isdigit() or 'day' in row[0].lower()):
                        continue
                
                if len(row) >= 2:  # Minimum required columns
                    # Pad row to expected columns if needed
                    while len(row) < expected_columns:
                        row.append("")
                    # Trim to expected columns if too many
                    row = row[:expected_columns]
                    data_rows.append(row)
                    print(f"✓ Added row: {row[0]} (total cells: {len(row)})")

        print(f"📊 Found {len(data_rows)} valid data rows for Excel export")

        if not data_rows:
            print("❌ No valid data rows found. Debugging info:")
            for i, line in enumerate(data_lines[:15]):  # Show first 15 lines
                print(f"  Line {i+1}: '{line}' (columns: {len(line.split('|'))})")
            raise ValueError("No valid data rows found in calendar text")

        # Use detected header or create default columns
        if header_row and len(header_row) > 0:
            columns = header_row
            print(f"📋 Using detected columns: {columns}")
        else:
            # Fallback to default columns based on expected_columns
            if include_transcripts:
                columns = [
                    "Day", "Reel Title", "Hook Script (0-2s)", "Body Breakdown (3-20s)",
                    "Close/CTA (20-30s)", "Format Style", "Audio Style", "Hashtag Strategy",
                    "Production Notes", "Optimization Tips", "Full Transcript"
                ]
            else:
                columns = [
                    "Day", "Reel Title", "Hook Script (0-2s)", "Body Breakdown (3-20s)",
                    "Close/CTA (20-30s)", "Format Style", "Audio Style", "Hashtag Strategy",
                    "Production Notes", "Optimization Tips"
                ]
            print(f"📋 Using default columns: {columns}")

        # Ensure we have the right number of columns
        while len(columns) < expected_columns:
            columns.append(f"Column_{len(columns)+1}")
        columns = columns[:expected_columns]

        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=columns)

        # Clean data
        df = df.fillna("")  # Replace NaN with empty strings
        df = df.replace("nan", "")  # Replace "nan" strings

        # Ensure output directory exists
        os.makedirs(os.path.dirname(filename), exist_ok=True)

        # Create Excel file with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Content Calendar"

        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths for detailed CEO-level content
        if include_transcripts:
            column_widths = [8, 30, 40, 50, 35, 18, 18, 25, 30, 30, 60]  # Added wider column for transcript
        else:
            column_widths = [8, 30, 40, 50, 35, 18, 18, 25, 30, 30]
        
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        # Format data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Save workbook
        wb.save(filename)
        print(f"✅ Excel file exported successfully: {filename}")
        print(f"📊 Total rows: {len(data_rows)}")

    except Exception as e:
        print(f"❌ Error exporting to Excel: {str(e)}")
        # Fallback to simple pandas export
        try:
            simple_df = pd.DataFrame([["Error in processing"]], columns=["Error"])
            simple_df.to_excel(filename, index=False)
            print(f"⚠️ Created minimal Excel file due to errors")
        except:
            raise ValueError(f"Failed to create Excel file: {str(e)}")